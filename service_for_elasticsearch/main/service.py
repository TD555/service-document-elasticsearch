import xml.etree.ElementTree as ET
import xmltodict
import uuid
from psycopg2.extras import RealDictCursor
import psycopg2 as ps
from flask import Flask, jsonify, request, abort
from elasticsearch import Elasticsearch, ConnectionError, BadRequestError
import asyncio
import aiohttp
from datetime import datetime
import requests
from collections import defaultdict
from random import random
import tempfile
import traceback
import time
import pytz
import openpyxl
import xlrd
import subprocess
import threading

import fitz
import os
import re
import io

from version import __version__, __description__

from keyword_extraction import keyword_extractor
try:
    from chat_with_graph import chat_with_AI
except Exception as e:
    ERROR_MESSAGE = str(e)

app = Flask(__name__)


# OLD_ES_INDEX = "araks_index_pre"
# ES_INDEX = "araks_index_v2"
# AMAZON_URL = "https://araks-projects-develop.s3.amazonaws.com/"
# ES_HOST = "http://localhost:9201/"

# DATABASE_HOST = 'localhost'
# # DATABASE_HOST = 'host.docker.internal'
# DATABASE_NAME = 'araks_db'
# DATABASE_USER = 'postgres'
# DATABASE_PASSWORD = 'Tik.555'
# DATABASE_PORT = 5433

OLD_ES_INDEX = os.environ['ELASTICSEARCH_INDEX']
ES_INDEX = os.environ['ELASTICSEARCH_NEW_INDEX']
AMAZON_URL = os.environ['AMAZON_URL']
ES_HOST = os.environ['ELASTICSEARCH_URL']

DATABASE_NAME = os.environ['DB_NAME']
DATABASE_USER = os.environ['DB_USER']
DATABASE_HOST = os.environ['DB_HOST']
DATABASE_PASSWORD = os.environ['DB_PASSWORD']
DATABASE_PORT = os.environ['DB_PORT']

es = Elasticsearch([ES_HOST])

request_timeout = 30
upload_timeout = 30

gmt_plus_4 = pytz.timezone("Asia/Dubai")


put_data = {
    "settings": {
        "analysis": {
            "analyzer": {
                "my_analyzer": {"tokenizer": "standard", "filter": ["lowercase"]}
            }
        }
    },
    "mappings": {
        "properties": {
            'property':
                {
                    'type': 'nested',
                    'properties': {
                        'name': {'type': 'text'},
                        'data_type': {'type': 'text'},
                        'data': {
                            'type': 'nested',
                            'properties': {
                                'content': {"type": "text", "analyzer": "my_analyzer"},
                                "created": {"type": "date", "format": "yyyy-MM-dd HH:mm:ss"},
                                "keywords": {
                                    'type': 'nested',
                                    'properties': {
                                        'name': {"type": "keyword"},
                                        'score': {'type': 'half_float'}
                                    }
                                },
                                "url": {
                                    "type": "keyword"
                                },
                            }
                        }
                    }
                }
        }
    },
}


try:
    es.indices.create(index=ES_INDEX, body=put_data)

except BadRequestError as e:
    pass


settings = {"highlight.max_analyzed_offset": 10000000}

try:
    es.indices.put_settings(index=ES_INDEX, settings=settings)
    OFFSET = 10000000
except BadRequestError as e:
    OFFSET = 1000000
    pass


async def get_filestorage_object(url):
    response = await asyncio.get_event_loop().run_in_executor(None, requests.get, url)
    if response.status_code == 200:
        file_object = io.BytesIO(response.content)
        return file_object
    else:
        return None
    # try:
    #     with open(url, 'rb') as file:
    #         file_content = file.read()
    #         file_object = io.BytesIO(file_content)
    #         return file_object
    # except FileNotFoundError:
    #     print("File not found.")
    #     return None


# def remove_duplicates(input_list):
#     seen = set()
#     unique_list = []

#     for item in input_list:
#         frozen_item = frozenset(item.items())

#         if frozen_item not in seen:
#             seen.add(frozen_item)
#             unique_list.append(item)

#     return unique_list


def check_base_url_exists(path):
    if not path.startswith(AMAZON_URL):
        return AMAZON_URL + path
    else:
        return path


async def extract_text_from_pdf(pdf_file):

    all_texts = []
    pdf_document = fitz.open("pdf", pdf_file.read())

    # Iterate through each page
    for page_num in range(pdf_document.page_count):
        page = pdf_document.load_page(page_num)

        # Extract text from the page
        text = page.get_text("text")

        clean_text = text.replace("\n", " ")
        clean_text = re.sub(r'\s+', ' ', clean_text)

        all_texts.append((text, clean_text))

    # Close the PDF document
    pdf_document.close()

    return all_texts


async def extract_text_from_doc(doc_file):

    doc_file.seek(0)

    document_content = io.BytesIO(doc_file.read())
    temp_file = tempfile.NamedTemporaryFile(suffix=".docx")

    temp_file.write(document_content.getvalue())
    document_path = temp_file.name

    pdf_bytes = subprocess.check_output(
        ["unoconv", "-f", "pdf", "--stdout", document_path]
    )

    # Create a BytesIO object from the PDF content
    pdf_stream = io.BytesIO(pdf_bytes)

    all_texts = await extract_text_from_pdf(pdf_stream)
    temp_file.close()

    return all_texts


async def extract_text_from_ppt(ppt_file):

    ppt_file.seek(0)

    ppt_content = io.BytesIO(ppt_file.read())
    temp_file = tempfile.NamedTemporaryFile(suffix=".pptx")
    temp_file.write(ppt_content.getvalue())
    ppt_path = temp_file.name

    pdf_bytes = subprocess.check_output(
        ["unoconv", "-f", "pdf", "--stdout", ppt_path]
    )
    pdf_stream = io.BytesIO(pdf_bytes)

    all_texts = await extract_text_from_pdf(pdf_stream)

    temp_file.close()

    return all_texts


async def extract_text_from_xlsx(xlsx_file):

    temp_buffer = io.BytesIO(xlsx_file.read())

    workbook = openpyxl.load_workbook(temp_buffer)

    all_texts = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        row_str = []
        # Loop through each row in the sheet
        for row in sheet.iter_rows(values_only=True):
            row_str.append(" ".join(map(str, row)))

        text = sheet_name + " " + " ".join(row_str).strip()
        clean_text = text.replace("\n", " ")
        clean_text = re.sub(r'\s+', ' ', clean_text)

        all_texts.append((text, clean_text))

    workbook.close()
    temp_buffer.close()

    return all_texts


async def extract_text_from_xls(xls_file):

    temp_buffer = io.BytesIO(xls_file.read())

    xls_workbook = xlrd.open_workbook(file_contents=temp_buffer.read())

    all_texts = []

    for sheet_name in xls_workbook.sheet_names():
        sheet = xls_workbook.sheet_by_name(sheet_name)

        rows = []

        # Loop through each row in the sheet
        for row_num in range(sheet.nrows):
            row = sheet.row_values(row_num)
            rows.append(" ".join(map(str, row)))

        all_texts.append(sheet_name + " " + " ".join(rows).strip())

        return all_texts


@app.errorhandler(Exception)
def handle_error(error):
    # Get the traceback
    error_traceback = traceback.format_exc()
    print(error_traceback)
    if hasattr(error, "code"):
        status_code = error.code
    else:
        status_code = 500
    return {"message": str(error).strip(), "status_code": status_code}, status_code


@app.after_request
def after_request(response):
    response.headers.set("Access-Control-Allow-Origin", "*")
    response.headers.set("Access-Control-Allow-Headers",
                         "Content-Type,Authorization")
    response.headers.set("Access-Control-Allow-Methods",
                         "GET,PUT,POST,DELETE,OPTIONS")
    return response


@app.route("/", methods=["GET"])
def info():
    return __description__


async def check_nested_field(path, nested_field, value):

    query = {
        "query": {
            "nested": {
                "path": path,
                "query": {
                    "bool": {
                        "must": [
                            {"exists": {"field": f"{path}.{nested_field}"}},
                            {"term": {f"{path}.{nested_field}.keyword": value}}
                        ]
                    }
                }
            }
        }
    }

    result = es.search(index=ES_INDEX, body=query)
    return result


async def update_docs(newData, oldData, node_id, property_id, property_name, data_type):
    update_query = {
        "script": {
            "source": """
                boolean propertyExists = false;
                for (int i = 0; i < ctx._source.property.size(); i++) {
                    if (ctx._source.property[i].id == params.propertyId) {
                        // Check if params.oldData is not empty before removing
                        if (!params.oldData.isEmpty()) {
                            for (int j = 0; j < params.oldData.size(); j++) {
                                ctx._source.property[i].data.removeIf(item -> item.url == params.oldData[j].url && item.name == params.oldData[j].name);
                            }
                        }
                        // Check if params.newData is not empty before adding
                        if (!params.newData.isEmpty()) {
                            for (int k = 0; k < params.newData.size(); k++) {
                                ctx._source.property[i].data.add(params.newData[k]);
                            }
                        }
                        if (ctx._source.property[i].data.isEmpty()) {
                            ctx._source.property.remove(i);
                        }
                        propertyExists = true;
                    }
                }

                // If property does not exist, create it
                if (!propertyExists) {
                    def newProperty = [
                        "id": params.propertyId,
                        "data": params.newData,
                        "name": params.propertyName,
                        "data_type": params.propertyType
                    ];
                    if (!ctx._source.containsKey('property')) {
                        ctx._source.property = [newProperty];
                    } else {
                        ctx._source.property.add(newProperty);
                    }
                }
            """,
            "params": {
                "propertyId": property_id,
                "propertyName": property_name,
                "propertyType": data_type,
                "newData": newData,
                "oldData": oldData  # Assuming newData is a list of dictionaries
            }
        },
        "query": {
            "bool": {
                "must": [
                    {
                        "term": {
                            "node_id.keyword": {
                                "value": node_id
                            }
                        }
                    }
                ]
            }
        }
    }

    es.update_by_query(index=ES_INDEX, body=update_query)


async def delete_docs(non_nested_fields):
    delete_query = {
        "query": {
            "bool": {
                "must": [
                ]
            }
        }
    }
    for field, value in non_nested_fields.items():
        if value:
            delete_query['query']['bool']['must'].append(
                {"term": {field + ".keyword": value}})

    return dict(es.delete_by_query(index=ES_INDEX, body=delete_query))


async def delete_empty_docs():
    delete_query = {
        "query": {
            "bool": {
                "must_not": {
                    "nested": {
                        "path": "property",
                        "query": {
                            "exists": {
                                "field": "property.id"
                            }
                        }
                    }
                }
            }
        }
    }

    es.delete_by_query(index=ES_INDEX, body=delete_query)


async def index_item(index, id, body):
    es.index(index=index, id=id, body=body)


@app.route("/create_or_update", methods=["POST"])
async def create_or_update():

    messages = []

    try:
        nodes_data = request.json["nodes_data"]
        project_id = request.json["project_id"]
        user_id = request.json["user_id"]
        node_id = request.json["node_id"]
        node_name = request.json["node_name"]
        type_id = request.json["type_id"]
        property_id = request.json["property_id"]
        data_type = request.json["data_type"]
        type_name = request.json["type_name"]
        property_name = request.json["property_name"]
        color = request.json["color"]
        default_image = request.json["default_image"]

    except:
        abort(400, "Invalid raw data")

    existing_document = es.get(index=ES_INDEX, id=node_id, ignore=404)

    if existing_document and existing_document["found"]:
        # Document with the same node_id and property_id exists, update it

        result = (await check_nested_field(path='property', nested_field='id', value=property_id))['hits']['hits']

        if result:
            data = ([item['data'] for item in result[0]['_source']
                    ['property'] if item['id'] == property_id][0])

        else:
            data = []

        my_data = [{k: item[k] for k in item.keys() if k in ['url', 'name']}
                   for item in data]

        amazon_nodes_data = [{'url': check_base_url_exists(
            item['url']), 'name': item['name']} for item in nodes_data]  # type: ignore
        non_repeat_nodes_data = [
            item for item in amazon_nodes_data if item not in my_data]
        non_repeat_data = [
            item for item in my_data if item not in amazon_nodes_data]  # type: ignore

        if non_repeat_nodes_data:
            await asyncio.gather(*[get_content(item) for item in non_repeat_nodes_data])

        if non_repeat_nodes_data or non_repeat_data:
            await update_docs(newData=[{i: item[i] for i in item if i != 'org_content'} for item in non_repeat_nodes_data], oldData=non_repeat_data, node_id=node_id, property_id=property_id, property_name=property_name, data_type=data_type)

        if non_repeat_data:
            es.indices.refresh(index=ES_INDEX)
            await delete_empty_docs()

        messages.append({
            'updated': [item['url'] for item in non_repeat_nodes_data],
            'deleted': [item['url'] for item in non_repeat_data]
        })

        thread = threading.Thread(target=update_keywords, kwargs={
            'items': [(check_base_url_exists(item['url']), item['org_content']) for item in [item_ for item_ in non_repeat_nodes_data if item_['content']]]})

        thread.start()

    elif nodes_data:

        messages.append({'updated': [check_base_url_exists(
            item['url']) for item in nodes_data], 'deleted': []})

        await asyncio.gather(*[get_content(item) for item in nodes_data])

        document_data = {**{'project_id': project_id, 'user_id': user_id, 'color': color, 'type_id': type_id, 'type_name': type_name, 'node_id': node_id,
                            'node_name': node_name, 'default_image': default_image}, **{'property': [{'id': property_id, 'name': property_name, 'data_type': data_type, 'data': [{i: item[i] for i in item if i != 'org_content'} for item in nodes_data]}]}}

        # Document does not exist, create a new one

        await index_item(index=ES_INDEX,
                         id=node_id,
                         body=document_data)

        thread = threading.Thread(target=update_keywords, kwargs={
            'items': [(check_base_url_exists(item['url']), item['org_content']) for item in [item_ for item_ in nodes_data if item_['content']]]})

        thread.start()

    return jsonify({"status": 200, 'message': messages})


def update_keywords(items, check_keywords=True):

    for url, text in items:

        es.indices.refresh(index=ES_INDEX)

        if text.strip():
            es.update_by_query(
                index=ES_INDEX,
                body={
                    "query": {
                        "nested": {
                            "path": "property.data",
                            "query": {
                                "bool": {
                                    "must": [
                                    ]
                                }
                            }
                        }
                    },
                    "script": {
                        "source": "for (int i = 0; i < ctx._source.property.size(); i++) {for (int j = 0; j < ctx._source.property[i].data.size(); j++) {if (ctx._source.property[i].data[j].url == params.path) {def newKeyword = params.keywords ; ctx._source.property[i].data[j].keywords = newKeyword;}}}",
                        "params": {
                            "path": url,
                            "keywords": keyword_extractor.extract(text, check_keywords=url.endswith('pdf'))
                        }
                    }
                }
            )
            es.indices.refresh(index=ES_INDEX)


async def get_time_now():
    current_utc_time = datetime.utcnow()
    gmt_plus_4_time = current_utc_time.replace(
        tzinfo=pytz.utc).astimezone(gmt_plus_4)

    return gmt_plus_4_time.strftime("%Y-%m-%d %H:%M:%S")


async def get_content(item):

    path = check_base_url_exists(item['url'])
    filename = os.path.basename(path)
    file = None

    # if not path.startswith(AMAZON_URL):
    #     path = AMAZON_URL + path
    try:
        start = time.time()
        file = await asyncio.wait_for(
            get_filestorage_object(path), timeout=request_timeout
        )
        end = time.time()
        request_time = end - start

        if not file:
            content = []

        else:
            try:
                if filename.endswith(".pdf"):
                    content = await asyncio.wait_for(
                        extract_text_from_pdf(
                            pdf_file=file), upload_timeout - request_time
                    )

                elif (
                    filename.endswith(".docx")
                    or filename.endswith(".doc")
                    or filename.endswith(".msword")
                    or filename.endswith(".document")
                ):
                    content = await asyncio.wait_for(
                        extract_text_from_doc(
                            doc_file=file), upload_timeout - request_time
                    )

                elif filename.endswith(".pptx") or filename.endswith(".ppt"):
                    content = await asyncio.wait_for(
                        extract_text_from_ppt(
                            ppt_file=file), upload_timeout - request_time
                    )

                elif filename.endswith(".xlsx"):
                    content = await asyncio.wait_for(
                        extract_text_from_xlsx(
                            xlsx_file=file), upload_timeout - request_time
                    )

                elif filename.endswith(".xls"):
                    content = await asyncio.wait_for(
                        extract_text_from_xls(
                            xls_file=file), upload_timeout - request_time
                    )

                else:
                    content = []

            except Exception:
                content = []

    except Exception:

        content = []

    finally:
        if file:
            file.close()

    item['url'] = path
    if content:
        item['org_content'] = " ".join([item[0] for item in content])
        item['content'] = " ".join([item[1]
                                   for item in content])  # type: ignore
    else:
        item['content'] = ""

    item['created'] = await get_time_now()
    item['keywords'] = []


async def update_fields(id_, id_value, fields_dict):
    update_query = {
        "script": {
            "source": f"""if (ctx._source.{id_} == '{id_value}')
                    {{for (int i = 0; i < params.fields_dict.size(); i++) {{ for (entry in params.fields_dict.entrySet())
                    {{String key = entry.getKey(); String value = entry.getValue(); if (ctx._source.containsKey(key) && ctx._source[key] != value) {{ctx._source[key] = value }}}}}}}}""",
            "lang": "painless",
            "params": {
                    "fields_dict": fields_dict
            }
        },
        "query": {"bool": {
            "must": [{"term": {
                f"{id_}.keyword": {
                    "value": id_value
                }
            }
            }
            ]
        }}
    }

    return es.update_by_query(index=ES_INDEX, body=update_query)


async def update_nested_field(id_value, fields_dict):
    update_query = {
        "script": {
            "source": f""" for (int i = 0; i < ctx._source.property.size(); i++) {{if (ctx._source.property[i].id == '{id_value}') {{ for (entry in params.fields_dict.entrySet())
                    {{String key = entry.getKey(); String value = entry.getValue(); if (ctx._source.property[i].containsKey(key) && ctx._source.property[i][key] != value) {{ctx._source.property[i][key] = value }}}}}}}}""",
            "lang": "painless",
            "params": {
                "fields_dict": fields_dict
            }
        },
        "query": {"bool": {"must": [{"nested": {"path": "property", "query": {"term": {"property.id.keyword": id_value}}}}]}}

    }

    return es.update_by_query(index=ES_INDEX, body=update_query)


@app.route("/update_type", methods=["POST"])
async def update_type():

    try:
        type_id = request.json.get("type_id")
        type_name = request.json.get("type_name")
        color = request.json.get("color")

    except:
        abort(400, "Invalid raw data")

    response = await update_fields(id_='type_id', id_value=type_id, fields_dict={"type_name": type_name, "color": color})

    if response['total']:
        message = "Type's fields were updated!"

    else:
        message = "There is no type with that ID"

    return jsonify({"message": message, "status": 200})


@app.route("/update_property", methods=["POST"])
async def update_property():

    try:
        property_id = request.json.get("property_id")
        property_name = request.json.get("property_name")
        data_type = request.json.get("data_type")

    except:
        abort(400, "Invalid raw data")

    response = await update_nested_field(id_value=property_id, fields_dict={"name": property_name, "data_type": data_type})

    if response['total']:
        message = "Property's fields were updated!"

    else:
        message = "There is no property with that ID"

    return jsonify({"message": message, "status": 200})


@app.route("/update_node", methods=["POST"])
async def update_node():

    try:
        node_id = request.json.get("node_id")
        node_name = request.json.get("node_name")
        default_image = request.json.get("default_image")

    except:
        abort(400, "Invalid raw data")

    response = await update_fields(id_='node_id', id_value=node_id, fields_dict={"node_name": node_name, "default_image": default_image})

    # response = await update_nested_field(id_=changing_id, id_value=changing_id_value, field_=changing_field, field_value=changing_field_value)

    if response['total']:
        message = "Node's fields were updated!"

    else:
        message = "There is no node with that ID"

    return jsonify({"message": message, "status": 200})


@app.route("/delete_node", methods=["DELETE"])
async def delete_node():

    try:
        project_id = request.json.get("project_id", None)
        node_id = request.json.get("node_id", None)
        property_id = request.json.get("property_id", None)

    except:
        abort(400, "Invalid raw data")

    non_nested_fields = {'project_id': project_id, 'node_id': node_id}

    if not property_id:
        response = await delete_docs(non_nested_fields=non_nested_fields)

    else:
        update_script = {
            "script": {
                "source": """
                    for (int i = 0; i < ctx._source.property.size(); i++) {
                        if (ctx._source.property[i].id == params.propertyId) {
                            ctx._source.property[i].data.clear();
                            ctx._source.property.remove(i);
                        }
                    }
                """,
                "params": {
                    "propertyId": property_id,
                }
            }
        }

        query = {
            "bool": {
                "must": [{"nested": {"path": "property", "query": {"term": {"property.id.keyword": property_id}}}}]}}

        for field, value in non_nested_fields.items():
            if value:
                query['bool']['must'].append(
                    {"term": {field + ".keyword": value}})

        response = es.update_by_query(index=ES_INDEX, body={
            "query": query,
            "script": update_script["script"]
        })

        es.indices.refresh(index=ES_INDEX)
        await delete_empty_docs()

    if response['total']:
        message = 'Documents were deleted!'

    else:
        message = 'There is no document with given conditions'

    return jsonify({"message": message, "status": 200})


def initialize_queries(keyword):
    query1 = {
        "_source": {
            "includes": ["user_id", "project_id", "color", "type_id", "type_name", "node_id", "node_name", "default_image"]
        },
        "query": {
            "nested": {
                "path": "property",
                "query": {
                    "bool": {
                        "should": [
                            {
                                "nested": {
                                    "path": "property.data",
                                    "query": {
                                        "bool": {
                                            "should": [
                                                {
                                                    "match": {
                                                        "property.data.content": {
                                                            "query": keyword.strip(),
                                                            "operator": "AND",
                                                            "fuzziness": "AUTO",
                                                            "analyzer": "my_analyzer"
                                                        }
                                                    }
                                                },
                                                {
                                                    "query_string": {
                                                        "query": "*" + keyword.strip() + "*",
                                                        "analyzer": "my_analyzer"
                                                    }
                                                }
                                            ]
                                        }
                                    },
                                    "inner_hits": {
                                        "name": "data_content",
                                        "highlight": {
                                            "pre_tags": ["<em>"],
                                            "post_tags": ["</em>"],
                                            "fields": {
                                                "property.data.content": {
                                                    "type": "plain",
                                                    "fragmenter": "span",
                                                    "number_of_fragments": 10000,
                                                    "order": "score",
                                                    "max_analyzed_offset": OFFSET
                                                }
                                            }
                                        }
                                    }
                                }
                            },
                            {
                                "nested": {
                                    "path": "property.data",
                                    "query": {
                                        "bool": {
                                            "should": [
                                                {
                                                    "match": {
                                                        "property.data.name": {
                                                            "query": keyword.strip(),
                                                            "operator": "AND",
                                                            "fuzziness": "AUTO",
                                                            "analyzer": "my_analyzer"
                                                        }
                                                    }
                                                },
                                                {
                                                    "query_string": {
                                                        "query": "*" + keyword.strip() + "*",
                                                        "analyzer": "my_analyzer"
                                                    }
                                                }
                                            ]
                                        }
                                    },
                                    "inner_hits": {
                                        "name": "data_name",
                                        "highlight": {
                                            "pre_tags": ["<em>"],
                                            "post_tags": ["</em>"],
                                            "fields": {
                                                "property.data.name": {
                                                    "type": "plain",
                                                    "fragmenter": "span",
                                                    "number_of_fragments": 1,
                                                    "order": "score",
                                                    "max_analyzed_offset": OFFSET
                                                }
                                            }
                                        }
                                    }
                                }
                            }
                        ]
                    }
                },
                "inner_hits": {}
            }
        }
    }

    query2 = {
        "_source": {
            "includes": ["user_id", "project_id", "color", "type_id", "type_name", "node_id", "node_name", "default_image"]
        },
        "query": {
            "nested": {
                "path": "property",
                "query": {
                    "bool": {
                        "should": [
                            {
                                "nested": {
                                    "path": "property.data",
                                    "query": {
                                        "span_near": {
                                            "clauses": [],
                                            "in_order": True
                                        }
                                    },
                                    "inner_hits": {
                                        "name": "data_content",
                                        "highlight": {
                                            "fields": {
                                                "property.data.content": {
                                                    "type": "plain",
                                                    "fragmenter": "span",
                                                    "number_of_fragments": 10000,
                                                    "order": "score",
                                                    "max_analyzed_offset": OFFSET
                                                }
                                            }
                                        }
                                    }
                                }
                            },
                            {
                                "nested": {
                                    "path": "property.data",
                                    "query": {
                                        "span_near": {
                                            "clauses": [],
                                            "in_order": True
                                        }
                                    },
                                    "inner_hits": {
                                        "name": "data_name",
                                        "highlight": {
                                            "fields": {
                                                "property.data.name": {
                                                    "type": "plain",
                                                    "fragmenter": "span",
                                                    "number_of_fragments": 0,
                                                    "order": "score",
                                                    "max_analyzed_offset": OFFSET
                                                }
                                            }
                                        }
                                    }
                                }
                            },
                            {
                                "nested": {
                                    "path": "property",
                                    "query": {
                                        "match_all": {}
                                    },
                                    "inner_hits": {
                                        "name": "property",
                                        "highlight": {
                                            "fields": {
                                                "property.name": {
                                                    "type": "plain",
                                                    "fragmenter": "span",
                                                    "number_of_fragments": 0,
                                                    "order": "score",
                                                    "max_analyzed_offset": 100000
                                                }
                                            }
                                        }
                                    }
                                }
                            }
                        ]
                    }
                },
                "inner_hits": {}
            }
        }
    }

    return query1, query2


def sentence_search(keywords, query, method, scroll_timeout, scroll_size):
    for splited_text in keywords:

        if method == "regexp":
            query["query"]["nested"]["query"]["bool"]["should"][0]["nested"]["query"]["span_near"]["clauses"].append(
                {
                    "span_multi": {
                        "match": {
                            "regexp": {
                                "property.data.content": {
                                    "value": f".*{splited_text.strip().lower()}.*",
                                    "flags": "ALL"
                                }
                            }
                        }
                    }
                }
            )
            query["query"]["nested"]["query"]["bool"]["should"][1]["nested"]["query"]["span_near"]["clauses"].append(
                {
                    "span_multi": {
                        "match": {
                            "regexp": {
                                "property.data.content": {
                                    "value": f".*{splited_text.strip().lower()}.*",
                                    "flags": "ALL"
                                }
                            }
                        }
                    }
                }
            )

        elif method == "fuzzy":
            query["query"]["nested"]["query"]["bool"]["should"][0]["nested"]["query"]["span_near"]["clauses"].append(
                {
                    "span_multi": {
                        "match": {
                            "fuzzy": {
                                "property.data.content": {
                                    "value": splited_text.strip().lower(),
                                    "fuzziness": "AUTO"
                                }
                            },
                        }
                    }
                }
            )
            query["query"]["nested"]["query"]["bool"]["should"][1]["nested"]["query"]["span_near"]["clauses"].append(
                {
                    "span_multi": {
                        "match": {
                            "fuzzy": {
                                "property.data.content": {
                                    "value": splited_text.strip().lower(),
                                    "fuzziness": "AUTO"
                                }
                            },
                        }
                    }
                }
            )
    try:
        result = es.search(
            index=ES_INDEX, body=query, scroll=scroll_timeout, size=scroll_size
        )

    except ConnectionError:
        abort(408, "Elasticsearch : Connection Timeout error")

    except Exception as e:
        abort(500, str(e))

    hits = result["hits"]["hits"]

    return result, hits


@app.route("/search", methods=["POST"])
def search():
    try:
        keyword = request.json["search"]
        page = request.json["page"]
        project_id = request.json["project_id"]
        list_type_id = request.json["list_type_id"]
        limit = request.json["limit"]
        sortOrder = request.json["sortOrder"]
        sortField = request.json["sortField"]

    except:
        abort(400, "Invalid raw data")

    if len(keyword.strip()) < 3:
        abort(400, "Search terms must contain at least 3 characters")

    scroll_size = limit  # Number of documents to retrieve in each scroll request
    scroll_timeout = "1m"  # Time interval to keep the search context alive
    special_characters = [
        "\\",
        "+",
        "-",
        "=",
        "&&",
        "||",
        ">",
        "<",
        "!",
        "(",
        ")",
        "{",
        "}",
        "[",
        "]",
        "^",
        '"',
        "~",
        "*",
        "?",
        ":",
        "/",
    ]

    for character in special_characters:
        keyword = keyword.replace(character, "\\" + character)

    keywords = keyword.strip().split()

    query1, query2 = initialize_queries(keyword)

    if not re.match(r".*[ +].*", keyword.strip()):
        try:
            result = es.search(
                index=ES_INDEX, body=query1, scroll=scroll_timeout, size=scroll_size
            )
            hits = result["hits"]["hits"]

        except ConnectionError:
            abort(408, "Elasticsearch : Connection Timeout error")

        except Exception as e:
            abort(500, str(e))

    else:
        result, hits = sentence_search(
            keywords, query2, "regexp", scroll_timeout, scroll_size)

        if not hits:
            try:
                query2["query"]['nested']["query"]["bool"]["should"][0]["nested"]["query"]["span_near"]["in_order"] = "false"
                result = es.search(
                    index=ES_INDEX, body=query2, scroll=scroll_timeout, size=scroll_size
                )
            except ConnectionError:
                abort(408, "Elasticsearch : Connection Timeout error")

            except Exception as e:
                abort(500, str(e))

            hits = result["hits"]["hits"]

            if not hits:

                query1, query2 = initialize_queries(keyword)

                result, hits = sentence_search(
                    keywords, query2, "fuzzy", scroll_timeout, scroll_size)

                if not hits:
                    try:
                        query2["query"]['nested']["query"]["bool"]["should"][0]["nested"]["query"]["span_near"]["in_order"] = "false"
                        result = es.search(
                            index=ES_INDEX, body=query2, scroll=scroll_timeout, size=scroll_size
                        )
                    except ConnectionError:
                        abort(408, "Elasticsearch : Connection Timeout error")

                    except Exception as e:
                        abort(500, str(e))

                    hits = result["hits"]["hits"]

                    if not hits:
                        try:
                            keyword = keyword.replace(" ", "")
                            query1["query"]['nested']['query']["bool"]["should"][0]['nested']['query']["bool"]["should"][0]["match"]["property.data.content"][
                                "query"
                            ] = keyword
                            query1["query"]['nested']['query']["bool"]["should"][0]['nested']['query']["bool"]["should"][1]["query_string"]["query"] = (
                                "*" + keyword + "*"
                            )
                            query1["query"]['nested']['query']["bool"]["should"][1]['nested']['query']["bool"]["should"][0]["match"]["property.data.name"][
                                "query"
                            ] = keyword
                            query1["query"]['nested']['query']["bool"]["should"][1]['nested']['query']["bool"]["should"][1]["query_string"]["query"] = (
                                "*" + keyword + "*"
                            )

                            result = es.search(
                                index=ES_INDEX,
                                body=query1,
                                scroll=scroll_timeout,
                                size=scroll_size,
                            )
                            hits = result["hits"]["hits"]
                        except ConnectionError:
                            abort(408, "Elasticsearch : Connection Timeout error")

                        except Exception as e:
                            abort(500, str(e))

    rows = []

    while hits:
        # Scroll to the next batch of results
        for hit in hits:
            if project_id != hit['_source']['project_id'] or not (hit["_source"]["type_id"] in list_type_id or not list_type_id):
                continue
            property_dict = {"user_id": hit['_source']['user_id'],
                             "project_id": hit['_source']['project_id'],
                             "type_id": hit['_source']['type_id'],
                             "type_name": hit['_source']['type_name'],
                             "color": hit['_source']['color'],
                             "default_image": hit['_source']['default_image'],
                             "node_id": hit['_source']['node_id'],
                             "node_name": hit['_source']['node_name']
                             }

            property_dict['data'] = []
            for property_hit in hit['inner_hits']['property']['hits']['hits']:

                property_dict['data_type'] = property_hit['_source']['data_type']
                # property_dict['property_id'] = property_hit['_source']['id']
                # property_dict['property_name'] = property_hit['_source']['name']

                for i, data_hit in enumerate(property_hit["inner_hits"]['data_content']['hits']['hits']):
                    data_dict = {}
                    data_dict["path"] = check_base_url_exists(
                        data_hit['_source']['url'])
                    data_dict["match_count"] = 0

                    if 'highlight' in data_hit:
                        data_dict["match_content"] = data_hit['highlight'].get(
                            'property.data.content')[0].strip()
                        for content in data_hit['highlight'].get(
                                'property.data.content'):
                            data_dict["match_count"] += int(
                                len(re.findall(r"<em>(.*?)</em>", content)))
                    else:
                        data_dict["match_content"] = ''

                    if property_hit["inner_hits"]['data_name']['hits']['hits']:
                        if 'highlight' in property_hit["inner_hits"]['data_name']['hits']['hits'][i]:
                            data_dict["match_filename"] = property_hit["inner_hits"]['data_name']['hits']['hits'][i]['highlight'].get(
                                'property.data.name', [''])[0].strip()
                        else:
                            data_dict["match_filename"] = ''
                    else:
                        data_dict["match_filename"] = ''

                    data_dict["created"] = data_hit['_source']['created']
                    data_dict["filename"] = data_hit['_source']['name']

                    property_dict['data'].append(data_dict)

            property_dict['updated'] = max(
                item['created'] for item in property_dict['data'])

            rows.append(property_dict.copy())

        scroll_id = result.get("_scroll_id")

        try:
            result = es.scroll(scroll_id=scroll_id, scroll=scroll_timeout)
        except Exception as e:
            abort(500, str(e))

        hits = result["hits"]["hits"]

    if sortOrder == "DESC" and sortField == "name":
        rows.sort(key=lambda x: x["node_name"], reverse=True)
    elif sortOrder == "DESC" and sortField == "updated_at":
        rows.sort(key=lambda x: x["updated"], reverse=True)
    elif sortOrder == "ASC" and sortField == "name":
        rows.sort(key=lambda x: x["node_name"])
    elif sortOrder == "ASC" and sortField == "updated_at":
        rows.sort(key=lambda x: x["updated"])
    else:
        abort(403, "Invalid sortOrder and/or sortField value")

    return jsonify(
        {
            "rows": rows[limit * (page - 1): limit * page],  # type: ignore
            "count": len(rows),
            "status": 200,
        }
    )


@app.route("/get_list", methods=["GET"])
async def get_list(**search):

    if not search:
        query = {"query": {"match_all": {}}, "size": 10000}

    else:
        query = {
            "query": {
                "bool": {
                    "must": [
                    ]
                }
            }, "size": 10000}

        for key, value in search.items():
            if value:
                query['query']['bool']['must'].append(
                    {"term": {key + '.keyword': value}})

    # Use the initial search API to retrieve the first batch of documents and the scroll ID
    try:
        while True:
            if es.indices.exists(index=ES_INDEX):
                initial_search = es.search(
                    index=ES_INDEX, body=query, scroll="1m")
                break
            else:
                continue

    except Exception as e:
        abort(500, str(e))

    scroll_id = initial_search["_scroll_id"]
    total_results = initial_search["hits"]["total"]["value"]

    # Iterate through the batches of results using the Scroll API
    documents = []
    while total_results > 0:
        for hit in initial_search["hits"]["hits"]:
            documents.append(hit["_source"])

        # Perform the next scroll request
        initial_search = es.scroll(scroll_id=scroll_id, scroll="1s")
        scroll_id = initial_search["_scroll_id"]
        total_results -= len(initial_search["hits"]["hits"])
        if len(initial_search["hits"]["hits"]) == 0:
            break

    # Clear the scroll context when done
    es.clear_scroll(scroll_id=scroll_id)

    return jsonify({"docs": documents, "status": 200})


@app.route("/clean", methods=["DELETE"])
async def clean():
    query = {"query": {"match_all": {}}}

    if es.delete_by_query(index=ES_INDEX, body=query, scroll_size=10000)["deleted"]:
        return jsonify(
            {
                "message": f"Elasticsearch database has cleaned successfully.",
                "status": 200,
            }
        )
    else:
        return jsonify(
            {"message": f"No document found in Elasticsearch database.", "status": 200}
        )


def sort_dict(my_dict: dict):
    return {
        k: v
        for k, v in sorted(my_dict.items(), key=lambda x: (-x[1], x[0]), reverse=False)
    }


def get_count(nodes: list, source_target: list = None) -> dict:  # type: ignore
    if not source_target:
        source_target = nodes
    return {node: source_target.count(node) for node in set(nodes)}


suggestions = [
    {
        "text": "There are {node} types of nodes and {edge} types of relations in the following graph scheme. Below are the types of nodes and connections:"
    },
    {
        "text": "Graph scheme with {node} types of nodes and {edge} types of relations. All node types and all connections are listed below:"
    },
    {
        "text": "The following graph scheme presents {node} types of nodes and {edge} types of relations. Here are all the node types and all their connections:"
    },
    {
        "text": "There are {node} types of nodes in the graph and {edge} kinds of connections. Here are everything you need to know about nodes and connections:"
    },
    {
        "text": "This graph contains {node} types of nodes and {edge} types of relationships. Here are the types of nodes and connections:"
    },
    {
        "text": "In the following graph scheme, there are {node} types of nodes and {edge} types of relations. Here are the types of nodes and all connections:"
    },
    {
        "text": "A graph scheme with {node} types of nodes and {edge} types of relations is shown below. The types of nodes and the types of connections are as follows:"
    },
    {
        "text": "This graph scheme contains {node} types of nodes and {edge} types of relationships. Here are the types of nodes and all connections:"
    },
    {
        "text": "This graph scheme presents {node} types of nodes and {edge} types of relations. Here are all the types of nodes and all the connections:"
    },
    {
        "text": "Graph scheme showing {node} types of nodes and {edge} types of relations. Here are the types of nodes and their connections:"
    },
    {
        "text": "This graph scheme has {node} types of nodes and {edge} types of relations. These are the types of nodes and the types of connections:"
    },
    {
        "text": "{edge} types of relations and {node} types of nodes are shown in the following graph scheme. The types of nodes and connections are as follows:"
    },
]


@app.route("/scheme_to_text", methods=["POST"])
async def get_scheme():
    # call to API and get data

    scheme_data = request.json["data"]

    if not scheme_data:
        input_sentence = "The following graph schema does not yet contain anything."

    relations = sort_dict(
        get_count(
            [
                data["source"] + " -> " +
                data["name"] + " -> " + data["target"]
                for data in scheme_data["edges"]
            ]
        )
    )
    scheme = [
        {
            "source": data["source"],
            "relation": data["name"],
            "target": data["target"],
        }
        for data in scheme_data["edges"]
    ]
    most_nodes = sort_dict(
        get_count(
            [item["source"] for item in scheme] + [item["target"]
                                                   for item in scheme]
        )
    )
    input_sentence = suggestions[int(random() * len(suggestions))]["text"].format(
        node=len(most_nodes), edge=len(set([item["relation"] for item in scheme]))
    )

    str_nodes = ""
    str_rels = ""
    for i, node in enumerate(most_nodes):
        str_nodes += str(i + 1) + "." + node + "\n"

    for j, rel in enumerate(relations):
        str_rels += str(j + 1) + "." + rel + "\n"

    input_sentence += f"\nNodes:\n{str_nodes}"
    input_sentence += f"\nRelations:\n{str_rels}"
    return jsonify({"text": input_sentence, "status": 200})


@app.route("/get__old_list", methods=["GET"])
async def get__old_list(**search):

    if not search:
        query = {"query": {"match_all": {}}, "size": 10000}

    else:
        query = {
            "query": {
                "bool": {
                    "must": [
                    ]
                }
            }, "size": 10000}

        for key, value in search.items():
            if value:
                query['query']['bool']['must'].append(
                    {"term": {key + '.keyword': value}})

    # Use the initial search API to retrieve the first batch of documents and the scroll ID
    try:
        while True:
            if es.indices.exists(index=OLD_ES_INDEX):
                initial_search = es.search(
                    index=OLD_ES_INDEX, body=query, scroll="1m")
                break
            else:
                continue

    except Exception as e:
        abort(500, str(e))

    scroll_id = initial_search["_scroll_id"]
    total_results = initial_search["hits"]["total"]["value"]

    # Iterate through the batches of results using the Scroll API
    documents = []
    while total_results > 0:
        for hit in initial_search["hits"]["hits"]:
            document = {
                "filename": hit["_source"]["filename"],
                "doc_id": hit["_source"]["doc_id"],
                "user_id": hit["_source"]["user_id"],
                "type_id": hit["_source"]["type_id"],
                "type_name": hit["_source"]["type_name"],
                "property_id": hit["_source"]["property_id"],
                "property_name": hit["_source"]["property_name"],
                "page": hit["_source"]["page"],
                "page_content": hit["_source"]["page_content"],
                "created": hit["_source"]["created"],
                "project_id": hit["_source"]["project_id"],
                "node_id": hit["_source"]["node_id"],
                "node_name": hit["_source"]["node_name"],
                "default_image": hit["_source"]["default_image"],
                "color": hit["_source"]["color"],
                "path": hit["_source"]["path"],
            }
            documents.append(document)

        # Perform the next scroll request
        initial_search = es.scroll(scroll_id=scroll_id, scroll="1s")
        scroll_id = initial_search["_scroll_id"]
        total_results -= len(initial_search["hits"]["hits"])
        if len(initial_search["hits"]["hits"]) == 0:
            break

    # Clear the scroll context when done
    es.clear_scroll(scroll_id=scroll_id)

    return jsonify({"docs": documents, "status": 200})


async def get_tags(url, project_id=None, node_id=None, property_id=None):
    query_body = {
        "query": {
            "bool": {
                "must": [
                    {
                        "term": {
                            "project_id.keyword": project_id
                        }
                    },
                    {
                        "nested": {
                            "path": "property.data",
                            "query": {
                                    "bool": {
                                        "must": [
                                            {
                                                "term": {
                                                    "property.data.url": url
                                                }
                                            }
                                        ]
                                    }
                            },
                            "inner_hits": {"_source": ["property.data.keywords"]}
                        }
                    }
                ]
            }
        }
    }

    # Optional filters (uncomment if needed)
    if project_id:
        query_body["query"]["bool"]["must"].append(
            {"term": {"project_id.keyword": project_id}})
    if node_id:
        query_body["query"]["bool"]["must"].append(
            {"term": {"node_id.keyword": node_id}})
    if property_id:
        query_body["query"]["bool"]["must"].append({
            "nested": {
                "path": "property",
                "query": {
                    "bool": {
                        "must": [
                            {"term": {"property.id.keyword": property_id}}
                        ]
                    }
                }
            }
        })

    response = es.search(index=ES_INDEX, body=query_body)

    return response


async def get_related_docs(keyword, url, project_id):
    return es.search(
        index=ES_INDEX,
        body={
            "_source": {
                "includes": ["user_id", "project_id", "color", "type_id", "type_name", "node_id", "node_name", "default_image"]
            },
            "query": {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "project_id.keyword": project_id
                            }
                        },
                        {
                            "nested": {
                                "inner_hits": {
                                    "_source": [
                                        "property.id",
                                        "property.name",
                                        "property.data_type"
                                    ]
                                },
                                "path": "property",
                                "query": {
                                    "nested": {
                                        "inner_hits": {
                                            "_source": [
                                                "property.data.url",
                                                "property.data.name",
                                                "property.data.created"
                                            ]
                                        },
                                        "path": "property.data",
                                        "query": {
                                            "bool": {
                                                "must": [
                                                    {
                                                        "nested": {
                                                            "path": "property.data.keywords",
                                                            "query": {
                                                                "bool": {
                                                                    "must": [
                                                                        {
                                                                            "term": {
                                                                                "property.data.keywords.name": keyword
                                                                            }
                                                                        }
                                                                    ]
                                                                }
                                                            }
                                                        }
                                                    }
                                                ],
                                                "must_not": [
                                                    {
                                                        "term": {
                                                            "property.data.url": url
                                                        }
                                                    }
                                                ]
                                            }
                                        }
                                    }
                                }
                            }
                        }
                    ]
                }
            }
        }
    )

from rapidfuzz import fuzz

TAXONOMY = ['Substance Sensing',
            'Nicotine Transportation',
            'Accessories',
            'Manufacturing',
            'Chewing',
            'Absorption',
            'Puff Control',
            'Mouth-Piece Filter',
            'Gaming',
            'Nicotine Dose',
            'Ritual',
            'Sensory Experience',
            'Size Coherence',
            'Storing',
            'Visual',
            'Capsules',
            'Connectivity',
            'Duration Of Experience',
            'Power Management And Charging',
            'Toxicology',
            'User Authentication',
            'Consumption Monitoring',
            'Dose',
            'Size',
            'Vending',
            'Properties',
            'Using',
            'Purity',
            'Hallow Acetate Tube',
            'Cleaning',
            'Customization',
            'Finishing',
            'Device',
            'Device Management',
            'Leakage Prevention',
            'Short Circuit Protection',
            'Ecosystems',
            'Aerosolization',
            'Mouth-End Paper',
            'Device Functional',
            'Taste',
            'Transportability',
            'Extinguisher',
            'Aerosol Control',
            'Composition',
            'Child Proof',
            'Airflow',
            'Outer Paper',
            'Filter',
            'Charging',
            'Substrate Hosting',
            'Shaking',
            'Wearable',
            'Heating',
            'Packaging',
            'Material',
            'Filters',
            'Ignition Propensity',
            'Overheat Protection',
            'Flavor',
            'Satisfaction',
            'Porous Matrix Material',
            'Safety',
            'Absorbability',
            'Shape Coherence',
            'Polymer Film Filter',
            'Haptic',
            'Events',
            'Action',
            'Heat Control',
            'Heating Profile',
            'Wrapped Paper',
            'Cartridge',
            'Consumable Detect',
            'Recognition',
            'Flavor',
            'Taste',
            'Container',
            'Storage']


@ app.route('/generate_tags', methods=["POST"])
async def generate_tags():
    try:
        project_id = request.json['project_id']
        # node_id = request.json['node_id']
        # property_id = request.json['property_id']
        url = check_base_url_exists(request.json['url'])
    except:
        abort(400, "Invalid raw data")

    # print(await get_tags(url))
    get_all_tags = await get_tags(url, project_id)
    if get_all_tags['hits']['hits']:
        result = get_all_tags['hits']['hits'][0]['inner_hits']["property.data"]['hits']['hits']
    else:
        return jsonify({'keywords': [], 'url': url, 'status': 200})
        abort(500, "There is no document with tags")

    if result:
        keywords = result[0]['_source']
        for i, keyword in enumerate(keywords['keywords']):
            result = (await get_related_docs(keyword['name'], url, project_id))['hits']['hits']
            keywords['keywords'][i]['taxonomy'] = [taxonomy for taxonomy in TAXONOMY  if fuzz.token_sort_ratio(taxonomy, keyword['name']) > 80]
            keywords['keywords'][i]['count'] = len(result)

        return jsonify(**keywords,  **{'url': url}, **{'status': 200})
    else:
        return jsonify({'keywords': [], 'url': url, 'status': 200})


@ app.route('/similar_docs', methods=["POST"])
async def get_similar_docs():
    try:
        project_id = request.json['project_id']
        # node_id = request.json['node_id']
        # property_id = request.json['property_id']
        url = check_base_url_exists(request.json['url'])
    except:
        abort(400, "Invalid raw data")

    get_all_tags = await get_tags(url, project_id)

    if get_all_tags['hits']['hits']:
        result = get_all_tags['hits']['hits'][0]['inner_hits']["property.data"]['hits']['hits']

        data = []

        all_dict = defaultdict(list)
        keywords = result[0]['_source']

        for i, keyword in enumerate(keywords['keywords']):
            related_docs = (await get_related_docs(keyword['name'], url, project_id))['hits']['hits']
            for node in related_docs:
                response_dict = {}
                node["_source"]['project_type_id'] = node["_source"].pop(
                    'type_id')
                response_dict.update(node["_source"])
                for property in node['inner_hits']["property"]['hits']['hits']:
                    all_dict[keyword['name']].extend([property_data["_source"]['url']
                                                      for property_data in property['inner_hits']["property.data"]['hits']['hits']])
                    response_dict.update(
                        {"property_" + k: v for k, v in property["_source"].items()})
                    for property_data in property['inner_hits']["property.data"]['hits']['hits']:
                        property_data["_source"]['document_name'] = property_data["_source"].pop(
                            'name')
                        response_dict.update(property_data["_source"])
                        if response_dict not in data:
                            data.append(response_dict.copy())
            keywords['keywords'][i]['count'] = len(related_docs)
            keywords['keywords'][i]['taxonomy'] = [taxonomy for taxonomy in TAXONOMY  if fuzz.token_sort_ratio(taxonomy, keyword['name']) > 80]

        all_urls = []

        for urls in all_dict.values():
            all_urls.extend(urls)

        return jsonify(**{'all_doc_count': len(all_urls)}, **{'data': data}, **{'keywords': keywords['keywords']}, **{'status': 200})
    else:
        return jsonify({'all_doc_count': 0, 'data': [], 'keywords': [], 'status': 200})


@ app.route('/expand_tag', methods=["POST"])
async def expand_tag():
    try:
        project_id = request.json['project_id']
        keyword = request.json['keyword']
        url = check_base_url_exists(request.json['url'])
    except:
        abort(400, "Invalid raw data")

    result = (await get_related_docs(keyword, url, project_id))['hits']['hits']

    data = []

    for node in result:
        response_dict = {}
        node["_source"]['project_type_id'] = node["_source"].pop('type_id')
        response_dict.update(node["_source"])
        for property in node['inner_hits']["property"]['hits']['hits']:
            response_dict.update(
                {"property_" + k: v for k, v in property["_source"].items()})
            for property_data in property['inner_hits']["property.data"]['hits']['hits']:
                property_data["_source"]['document_name'] = property_data["_source"].pop(
                    'name')
                response_dict.update(property_data["_source"])
                data.append(response_dict.copy())

    return jsonify({'data': data, 'taxonomy' : [taxonomy for taxonomy in TAXONOMY  if fuzz.token_sort_ratio(taxonomy, keyword) > 80], 'status': 200})


namespace = uuid.NAMESPACE_DNS

SEARCH_URL = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&term={keyword}&retmode=json&retmax={limit}&retstart={offset}&api_key=0587a88dd5dcd5fab78a8bf379542622b509'
FETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&id={id}&rettype=medline&retmode=xml&api_key=0587a88dd5dcd5fab78a8bf379542622b509"


def convert_date(pubDate):
    months = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"
    }
    day = pubDate.get("Day", '01')
    month = months.get(pubDate.get("Month", '01'), '01')
    year = pubDate["Year"]
    return f"{year}-{month}-{day}"


def convert_to_text(item, abstract=False):
    if isinstance(item, dict):
        return item.get("#text", "")
    if abstract and isinstance(item, list):
        text = '\n\n'.join([element["@Label"] + '\n' +
                           element["#text"] for element in item])
        return text
    return item


async def fetch_with_retry(session, id, retry_attempts=10, timeout=10):
    for attempt in range(retry_attempts):
        try:
            async with session.get(FETCH_URL.format(id=id.strip()), timeout=timeout) as response:
                response.raise_for_status()
                xml_file = await response.text()
                xml_data = ET.fromstring(xml_file)
                xmlstr = ET.tostring(xml_data, encoding='utf-8', method='xml')
                dict_data = dict(xmltodict.parse(xmlstr))[
                    'PubmedArticleSet']['PubmedArticle']['MedlineCitation']

                title = convert_to_text(dict_data['Article']['ArticleTitle'])
                id_data = {
                    'article': {
                        'article_id': id,
                        'name': title[:50],
                        'article_url': '',
                        'source': 'PubMed',
                        'title': title,
                        'abstract': convert_to_text(dict_data['Article'].get('Abstract', {'AbstractText': ''})["AbstractText"], True),
                        'pub_date': convert_date(dict_data['Article']['Journal']['JournalIssue']['PubDate']),
                        'language': dict_data['Article'].get('Language', '')
                    },
                    'country': dict_data['MedlineJournalInfo'].get('Country', ''),
                }

                elocation = dict_data['Article'].get('ELocationID')
                if isinstance(elocation, dict) and elocation.get('@EIdType') == 'doi':
                    id_data['article']['article_url'] = 'https://www.doi.org/' + \
                        elocation.get('#text', '')
                elif isinstance(elocation, list):
                    for eid in elocation:
                        if eid.get('@EIdType') == 'doi':
                            id_data['article']['article_url'] = 'https://www.doi.org/' + \
                                eid.get('#text', '')

                authors = dict_data['Article'].get(
                    'AuthorList', {'Author': []})['Author']
                if isinstance(authors, list):
                    authors = authors[:20]
                else:
                    authors = [authors]
                id_data['authors'] = [
                    {
                        "affiliation": author.get('AffiliationInfo', {"Affiliation": ""})["Affiliation"],
                        'name': (author.get('ForeName', '') + ' ' + author.get('LastName', '')).strip()[:50],
                        'author_id': uuid.uuid5(namespace, (author.get('ForeName', '') + ' ' + author.get('LastName', '') + ' ' + author.get('AffiliationInfo', {"Affiliation": ""})["Affiliation"]).strip())
                    }
                    if isinstance(author.get('AffiliationInfo'), dict)
                    else
                    {
                        "affiliation": ' '.join([item["Affiliation"] for item in author.get('AffiliationInfo', [{"Affiliation": ""}])]),
                        'name': (author.get('ForeName', '') + ' ' + author.get('LastName', '')).strip()[:50],
                        'author_id': uuid.uuid5(namespace, (author.get('ForeName', '') + ' ' + author.get('LastName', '') + ' ' + ' '.join([item["Affiliation"] for item in author.get('AffiliationInfo', [{"Affiliation": ""}])])).strip())
                    }
                    for author in authors if (author.get('ForeName', '') + ' ' + author.get('LastName', '')).strip()
                ]

                keywords = dict_data.get('KeywordList', {}).get('Keyword', [])
                if not isinstance(keywords, list):
                    keywords = [keywords]
                id_data['keywords'] = [keyword.get(
                    "#text", "") for keyword in keywords] if keywords else []

                return id_data

        except (aiohttp.ClientError, asyncio.TimeoutError) as ce:
            print(f"Client error for ID {id}: {ce}")
            if attempt < retry_attempts - 1:
                await asyncio.sleep(2 ** attempt)
                continue
            else:
                abort(500, f"Failed to fetch details for ID {id}: {ce}")
        except Exception as e:
            print(f"Error processing ID {id}: {e}")
            break
            abort(500, f"Error processing ID {id}: {e}")


async def search_with_retry(session, keyword, limit, offset, retry_attempts=10, timeout=10):
    for attempt in range(retry_attempts):
        try:
            async with session.get(SEARCH_URL.format(keyword=keyword, limit=limit, offset=offset), timeout=timeout) as response:
                response.raise_for_status()
                return await response.json()
        except (aiohttp.ClientError, asyncio.TimeoutError) as ce:
            print(f"Client error for keyword {keyword}: {ce}")
            if attempt < retry_attempts - 1:
                await asyncio.sleep(2 ** attempt)
                continue
            else:
                abort(500, f"Failed to search for keyword {keyword}: {ce}")
        except Exception as e:
            print(f"Error processing keyword {keyword}: {e}")
            abort(500, f"Error processing keyword {keyword}: {e}")


@ app.route('/pubmed/get_data', methods=["POST"])
async def pubmed_preview():
    try:
        keyword = str(request.json['keyword'])
        limit = request.json.get('limit', 5)
        page = int(request.json.get('page', 1))
    except Exception as e:
        abort(
            400, f"Invalid raw data. One of the parameters is incorrect. {str(e)}")

    if len(keyword) < 3:
        abort(400, "Keyword must be at least 3 characters long")

    try:
        async with aiohttp.ClientSession() as session:
            search_result = await asyncio.wait_for(search_with_retry(session, keyword, limit, (page-1)*limit), timeout=10)
            id_list = search_result['esearchresult']['idlist']
            await asyncio.sleep(0.1)

            tasks = [fetch_with_retry(session, id) for id in id_list]
            all_data = await asyncio.gather(*[asyncio.wait_for(task, timeout=20) for task in tasks])

        all_data = [data for data in all_data if data]

        return jsonify({'count': search_result['esearchresult']['count'], 'articles': all_data})
    except asyncio.TimeoutError:
        abort(500, "The request timed out. Please try again later.")


get_all_query = """SELECT np.user_id, np.project_id, n.project_type_id as type_id, pnt.name as type_name, pnt.color as color, np.node_id, n.name as node_name, n.default_image, np.project_type_property_id as property_id, pntp.name as property_name, 'doc' AS type, np.nodes_data, np.created_at as created
FROM public.node_properties as np
LEFT JOIN nodes as n
ON n.id = node_id
LEFT JOIN project_node_types as pnt
ON n.project_type_id = pnt.id
LEFT JOIN projects_node_type_property as pntp
ON np.project_type_property_id = pntp.id
WHERE project_type_property_type='document' and nodes_data not in ('[]', '[{}]');"""


@ app.route("/migrate", methods=["GET"])
async def migration():

    conn = ps.connect(dbname=DATABASE_NAME, user=DATABASE_USER,
                      password=DATABASE_PASSWORD, port=DATABASE_PORT, host=DATABASE_HOST)
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(get_all_query)

    all_data = cur.fetchall()

    old_data = [{k: v for k, v in data.items()} for data in all_data]
    try:
        new_index_list = []

        for item in old_data:
            node_found = False
            property_found = False
            # data_found = False

            for i, new_item in enumerate(new_index_list):
                if item["node_id"] == new_item['node_id']:
                    node_found = True
                    for j, prop_item in enumerate(new_index_list[i]['property']):
                        if item["property_id"] == prop_item['id']:
                            property_found = True
                            # print(0)

                    if not property_found:
                        datas = item['nodes_data']
                        await asyncio.gather(*[get_content(data) for data in datas])
                        for data in datas:
                            data['created'] = str(
                                item['created']).split('.')[0]
                        new_property = {
                            "data": datas,
                            # "data": [
                            #     {
                            #         "content": item["page_content"],
                            #         "created": item["created"],
                            #         "name": item["filename"],
                            #         "url": item["path"]
                            #     }
                            # ],
                            "data_type": "document",
                            "id": item["property_id"],
                            "name": item["property_name"]
                        }

                        new_index_list[i]['property'].append(new_property)

            if not node_found:
                new_index = {}  # Create a new dictionary for each iteration
                new_index["color"] = item["color"]
                new_index["default_image"] = item["default_image"]
                new_index["node_id"] = item["node_id"]
                new_index["node_name"] = item["node_name"]
                new_index["project_id"] = item["project_id"]
                new_index["type_id"] = item["type_id"]
                new_index["type_name"] = item["type_name"]
                new_index["user_id"] = item['user_id']
                datas = item['nodes_data']
                await asyncio.gather(*[get_content(data) for data in datas])
                for data in datas:
                    data['created'] = str(item['created']).split('.')[0]
                new_index["property"] = [
                    {
                        "data": datas,
                        # "data": [
                        #     {
                        #         "content": item["page_content"],
                        #         "created": item["created"],
                        #         "name": item["filename"],
                        #         "url": item["path"]
                        #     }
                        # ],
                        "data_type": "document",
                        "id": item["property_id"],
                        "name": item["property_name"]
                    }
                ]

                new_index_list.append(new_index)

        tasks = []

        for item in new_index_list:
            node_id = item["node_id"]
            index_name = ES_INDEX
            body = item

            # Create a task for each index call
            task = index_item(index_name, node_id, body)
            tasks.append(task)

        await asyncio.gather(*tasks)

        items = []

        for i in range(len(new_index_list)):
            for j in range(len(new_index_list[i]['property'])):
                for k in range(len(new_index_list[i]['property'][j]['data'])):
                    items.append((new_index_list[i]['property'][j]['data'][k]['url'],
                                 new_index_list[i]['property'][j]['data'][k]['content']))

        thread = threading.Thread(target=update_keywords, kwargs={
            'items': [item for item in [item_ for item_ in items if item_[1]]], 'check_keywords': False})

        thread.start()

        return jsonify({"status": 200})

    except Exception as e:
        abort(500, str(e))


#   AI tools

@ app.route('/chat', methods=['POST'])
async def chat():

    user_input: str = request.json.get("message")
    project_id: str = request.json.get("project_id")

    if user_input.lower().strip() in ['hi', 'hello', 'hey']:
        return jsonify({"response": 'Hi, how can I help with getting information about your graph?'})
    if not user_input:
        abort(400, "Invalid raw data. No message provided")

    try:
        response = await chat_with_AI.get_bot_response(user_input, project_id)
        return jsonify({"response": response})
    except Exception as e:
        if 'ERROR_MESSAGE' in globals():
            abort(503, ValueError(ERROR_MESSAGE))
        abort(503, ValueError(
            f'Cannot resolve address {os.environ["NEO4JURL"]}'))
